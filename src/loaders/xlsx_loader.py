"""XLSX loader — one DocElement(heading + table) per sheet."""
from __future__ import annotations

from pathlib import Path

from .elements import DocElement, elements_to_markdown, rows_to_markdown


def load_xlsx_elements(path: Path, *, max_rows: int = 500) -> list[DocElement]:
    try:
        from openpyxl import load_workbook
    except ImportError as e:
        raise RuntimeError("openpyxl not installed. Run: pip install openpyxl") from e

    wb = load_workbook(str(path), read_only=True, data_only=True)
    elements: list[DocElement] = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows: list[list[str]] = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i >= max_rows:
                break
            rows.append(["" if c is None else str(c) for c in row])
        if not rows:
            continue
        elements.append(DocElement(kind="heading", content=f"Sheet: {sheet_name}", meta={"level": 2}))
        md = rows_to_markdown(rows)
        elements.append(DocElement(kind="table", content=md, meta={"sheet": sheet_name}))
    wb.close()
    return elements


def load_xlsx(path: Path) -> str:
    return elements_to_markdown(load_xlsx_elements(path))
